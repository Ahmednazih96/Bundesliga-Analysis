import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import matplotlib.pyplot as plt
import seaborn as sns
from plotly.subplots import make_subplots
import plotly.graph_objects as go


file_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\merged_team_data.xlsx"
df = pd.read_excel(file_path)
print("Dataset loaded successfully. Columns:", df.columns)

df_possession=df
missing_values = df.isnull().sum()

output_value_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\Finaly_Team_EDA96.xlsx"

top_defensive_teams = df.sort_values('Goals Conceded per Match').head(10)
top_offensive_teams = df.sort_values('Goals per Match', ascending=False).head(10)
df['Goal Difference'] = df['Goals per Match'] - df['Goals Conceded per Match']
shot_conversion_vs_chances = df[['Team', 'Shot Conversion Rate (%)', 'Big Chances Missed', 'Country']]
df['Goal Difference'] = df['Goals per Match'] - df['Goals Conceded per Match']
possession_data = df[['Team', 'Possession (%)', 'Big Chances Missed']]
df_possession['Possession Category'] = pd.cut(df_possession['Possession (%)'], bins=[0, 45, 55, 65], labels=['Low', 'Medium', 'High'])
crosses_data = df[['Team', 'Accurate Crosses per Match', 'Cross Success (%)']]
clean_sheets_data = df[['Team', 'Clean Sheets']]
penalties_data = df[['Team', 'Penalties Won', 'Penalties Conceded']]
interceptions_clearances_data = df[['Team', 'Interceptions per Match', 'Clearances per Match']]
expected_goals_data = df[['Team', 'Expected Goals', 'Goals per Match','Total Goals']]
free_kicks_data = df[['Team', 'Corners Taken']]

df['Total_Crosses']=df['Accurate Crosses per Match']*df['Matches']
Total_Accurate_Crosses=df.groupby('Team')['Total_Crosses'].sum().sort_values(ascending=False)

df['Total_Shots']=df['Shots on Target per Match']*df['Matches']
Total_Shots_on_Target=df.groupby('Team')['Total_Shots'].sum().sort_values(ascending=False)

df['Total_Fouls']=df['Fouls per Match']*df['Matches']
Total_Fouls_Team=df.groupby('Team')['Total_Fouls'].sum().sort_values(ascending=False)

df['Total_Passes']=df['Accurate Passes per Match']*df['Matches']
Total_Passes_Accurate=df.groupby('Team')['Total_Passes'].sum().sort_values(ascending=False)


df['Total_Long_Balls']=df['Accurate Long Balls per Match']*df['Matches']
Total_Accurate_Long_Balls=df.groupby('Team')['Total_Long_Balls'].sum().sort_values(ascending=False)


#C:\Users\user\OneDrive\Data Curation Progect File\Team_AhmedNazih.py




# اختيار المقاييس المطلوبة فقط
metrics_columns = ["Goals per Match", "Expected Goals", "Shot Accuracy", "Big Chances", "Shot Conversion Rate (%)"]

# تصفية البيانات للحصول على الفرق التي تحتوي على هذه القيم
filtered_data = df[["Team"] + metrics_columns].dropna()

# تحويل البيانات إلى الشكل المناسب للخريطة الحرارية
heatmap_data = filtered_data.set_index("Team")

# رسم الخريطة الحرارية باستخدام Seaborn
plt.figure(figsize=(12, 8))
sns.heatmap(heatmap_data, annot=True, cmap="coolwarm", linewidths=0.5, fmt=".2f")

# تعديل العنوان والمحاور
plt.title("🔥 Heatmap of Attacking Metrics for All Teams 🔥", fontsize=14)
plt.xlabel("Metrics")
plt.ylabel("Teams")
plt.xticks(rotation=30)

# عرض المخطط
plt.show()









# df['Performance Category'] = pd.cut(df['FotMob Team Rating'], bins=[2, 4, 7, 10], 
# labels=['Low', 'Medium', 'High'])
# print(df['Performance Category'].value_counts())

# numeric_columns = df.select_dtypes(include=np.number).columns

# if os.path.exists(output_value_path):
#     os.remove(output_value_path)

# with pd.ExcelWriter(output_value_path, engine='openpyxl') as writer:
#     df.to_excel(writer, sheet_name='Selected Data', index=False)
#     missing_values[missing_values > 0].to_frame(name='Missing Values').to_excel(writer, sheet_name='Missing Values')
#     df.describe().to_excel(writer, sheet_name='Descriptive Stats')
#     df['Performance Category'].value_counts().to_frame(name='Performance Counts').to_excel(writer, sheet_name='Performance Counts')
#     top_offensive_teams.to_excel(writer, sheet_name='Top Offensive Teams')
#     top_defensive_teams.to_excel(writer, sheet_name='Top Defensive Teams')
#     shot_conversion_vs_chances.to_excel(writer, sheet_name='Shot Conversion vs Big Chances Missed', index=False)  
#     df[['Team', 'Goal Difference']].to_excel(writer, sheet_name='Goal Difference', index=False)
#     possession_data.to_excel(writer, sheet_name='Possession and Big Chances', index=False)
#     df_possession['Possession Category'].value_counts().to_excel(writer, sheet_name='Possession Categories')
#     crosses_data.to_excel(writer, sheet_name='Accurate Crosses and Success', index=False)
#     clean_sheets_data.to_excel(writer, sheet_name='Clean Sheets by Team', index=False)
#     penalties_data.to_excel(writer, sheet_name='Penalties Won vs Conceded', index=False)
#     interceptions_clearances_data.to_excel(writer, sheet_name='Interceptions vs Clearances', index=False)
#     expected_goals_data.to_excel(writer, sheet_name='Expected Goals vs Goals', index=False)
#     free_kicks_data.to_excel(writer, sheet_name='Free Kicks Data', index=False)
#     Total_Accurate_Crosses.to_excel(writer,sheet_name="Total Accurate Crosses")
#     Total_Shots_on_Target.to_excel(writer,sheet_name="Total Shots on Target")
#     Total_Fouls_Team.to_excel(writer,sheet_name=" Total Fouls Team")
#     Total_Passes_Accurate.to_excel(writer,sheet_name="Total Passes Accurate")
#     Total_Accurate_Long_Balls.to_excel(writer,sheet_name="Total Accurate Long Balls")
   










# plot_paths = []
# for col in numeric_columns:
#     plt.figure(figsize=(8, 4))
#     sns.histplot(df[col], kde=True, bins=20)
#     plt.title(f"Distribution of {col}")
#     plt.tight_layout()

#     plot_path = f"C:\\Users\\user\\OneDrive\\Data Curation Progect File\\Team\\distribution_{col}.png"
#     plt.savefig(plot_path)
#     plot_paths.append(plot_path)
#     plt.close()

# correlation_matrix = df[numeric_columns].corr()
# plt.figure(figsize=(12, 8))
# sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt=".2f")
# plt.title("Correlation Matrix")
# plt.tight_layout()
# plot_path_corr = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\correlation_matrix.png"
# plt.savefig(plot_path_corr)
# plt.close()

# plt.figure(figsize=(8, 5))
# sns.scatterplot(data=df, x='Rank', y='Goals per Match', hue='Country', palette='tab10')
# plt.title("Rank vs Goals per Match")
# plt.tight_layout()
# plot_path_scatter = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\scatter_rank_goals.png"
# plt.savefig(plot_path_scatter)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.barplot(data=top_offensive_teams, x='Goals per Match', y='Team', palette='viridis')
# plt.title("Top 10 Teams by Goals per Match")
# plt.tight_layout()
# plot_path_offensive = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\top_offensive_teams.png"
# plt.savefig(plot_path_offensive)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.barplot(data=top_defensive_teams, x='Goals Conceded per Match', y='Team', palette='coolwarm')
# plt.title("Top 10 Teams by Defensive Strength")
# plt.tight_layout()
# plot_path_defensive = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\top_defensive_teams.png"
# plt.savefig(plot_path_defensive)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.barplot(data=df, x='Goal Difference', y='Team', palette='coolwarm')
# plt.title("Goal Difference by Team")
# plt.tight_layout()
# plot_path_goal_diff = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\goal_difference.png"
# plt.savefig(plot_path_goal_diff)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.scatterplot(data=df, x='Shot Conversion Rate (%)', y='Big Chances Missed', hue='Country', palette='tab10')
# plt.title("Shot Conversion Rate vs Big Chances Missed")
# plt.tight_layout()
# plot_path_shot_conversion = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\shot_conversion_vs_big_chances.png"
# plt.savefig(plot_path_shot_conversion)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.barplot(data=df, x='Possession (%)', y='Team', palette='Blues')
# plt.title("Possession Percentage by Team")
# plt.tight_layout()
# plot_path_possession = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\possession_percentage.png"
# plt.savefig(plot_path_possession)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.scatterplot(data=df_possession, x='Possession (%)', y='Shot Conversion Rate (%)', hue='Country', palette='tab10')
# plt.title("Possession vs Shot Conversion Rate")
# plt.tight_layout()
# plot_path_possession_shot_conversion = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\possession_vs_shot_conversion.png"
# plt.savefig(plot_path_possession_shot_conversion)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.scatterplot(data=df_possession, x='Possession (%)', y='Big Chances Missed', hue='Country', palette='tab10')
# plt.title("Possession vs Big Chances Missed")
# plt.tight_layout()
# plot_path_possession_big_chances = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\possession_vs_big_chances.png"
# plt.savefig(plot_path_possession_big_chances)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.countplot(data=df_possession, x='Possession Category', palette='Set2')
# plt.title("Distribution of Teams by Possession Category")
# plt.tight_layout()
# plot_path_possession_category = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\possession_category_distribution.png"
# plt.savefig(plot_path_possession_category)
# plt.close()


# plt.figure(figsize=(10, 6))
# sns.scatterplot(data=df, x='Accurate Crosses per Match', y='Cross Success (%)', hue='Team', palette='deep')
# plt.title("Accurate Crosses per Match vs Cross Success (%)")
# plt.xlabel('Accurate Crosses per Match')
# plt.ylabel('Cross Success (%)')
# plt.tight_layout()
# plot_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\accurate_crosses_vs_success.png"
# plt.savefig(plot_path)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.barplot(data=df, x='Clean Sheets', y='Team', palette='Greens')
# plt.title("Clean Sheets by Team")
# plt.xlabel('Clean Sheets')
# plt.ylabel('Team')
# plt.tight_layout()
# plot_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\clean_sheets_by_team.png"
# plt.savefig(plot_path)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.scatterplot(data=df, x='Penalties Won', y='Penalties Conceded', hue='Team', palette='tab20')
# plt.title("Penalties Won vs Penalties Conceded")
# plt.xlabel('Penalties Won')
# plt.ylabel('Penalties Conceded')
# plt.tight_layout()
# plot_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\penalties_won_vs_conceded.png"
# plt.savefig(plot_path)
# plt.close()

# file_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\merged_team_data.xlsx"
# df = pd.read_excel(file_path)
# print("Dataset loaded successfully. Columns:", df.columns)
# plt.figure(figsize=(10, 6))
# sns.scatterplot(data=df, x='Interceptions per Match', y='Clearances per Match', hue='Team', palette='coolwarm')
# plt.title("Interceptions vs Clearances per Match")
# plt.xlabel('Interceptions per Match')
# plt.ylabel('Clearances per Match')
# plt.tight_layout()
# plot_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\interceptions_vs_clearances.png"
# plt.savefig(plot_path)
# plt.close()

# plt.figure(figsize=(10, 6))
# sns.scatterplot(data=df, x='Expected Goals', y='Goals per Match', hue='Team', palette='inferno')
# plt.title("Expected Goals vs Goals per Match")
# plt.xlabel('Expected Goals')
# plt.ylabel('Goals per Match')
# plt.tight_layout()
# plot_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\expected_goals_vs_goals.png"
# plt.savefig(plot_path)
# plt.close()


# plt.figure(figsize=(10, 6))
# sns.barplot(data=df, x='Corners Taken', y='Team', palette='Spectral')
# plt.title("Top Teams by Free Kicks Scored")
# plt.tight_layout()
# plot_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Team\free_kicks_scored.png"
# plt.savefig(plot_path)
# plt.close()

# wb = load_workbook(output_value_path)

# img_corr = Image(plot_path_corr)
# sheet = wb.create_sheet('Visualizations')

# sheet2 = wb.create_sheet('Visualizations numeric columns')

# sheet3 = wb.create_sheet(' Conversion ')


# sheet.add_image(img_corr, 'H5')  # 

# img_scatter = Image(plot_path_scatter)
# sheet.add_image(img_scatter, 'H25')  #

# img_offensive = Image(plot_path_offensive)
# sheet.add_image(img_offensive, 'H45')  #    

# img_defensive = Image(plot_path_defensive)
# sheet.add_image(img_defensive, 'H65')  #    

# for idx, col in enumerate(numeric_columns):
#     img = Image(plot_paths[idx])
#     sheet2.add_image(img, f'H{2 + idx * 15}')  

# img_goal_diff = Image(plot_path_goal_diff)
# sheet.add_image(img_goal_diff, 'H85')

# img_plot_path_shot_conversion = Image(plot_path_shot_conversion)
# sheet3.add_image(img_plot_path_shot_conversion, 'E2')

# img_shot_conversion = Image(plot_path_possession_shot_conversion)
# sheet3.add_image(img_shot_conversion, 'E40')
# img_big_chances = Image(plot_path_possession_big_chances)

# sheet3.add_image(img_big_chances, 'E80')
# img_category_dist = Image(plot_path_possession_category)
# sheet3.add_image(img_category_dist, 'E110')

# sheet = writer.sheets['Accurate Crosses and Success']
# img_crosses_vs_success = Image(plot_path)
# sheet.add_image(img_crosses_vs_success, 'F5')
# wb.save(output_value_path)


# sheet = writer.sheets['Clean Sheets by Team']
# img_clean_sheets = Image(plot_path)
# sheet.add_image(img_clean_sheets, 'F5')
# wb.save(output_value_path)


# sheet = writer.sheets['Penalties Won vs Conceded']
# img_penalties = Image(plot_path)
# sheet.add_image(img_penalties, 'F5')
# wb.save(output_value_path)


# sheet = writer.sheets['Interceptions vs Clearances']
# img_interceptions_clearances = Image(plot_path)
# sheet.add_image(img_interceptions_clearances, 'F5')
# wb.save(output_value_path)



# sheet = writer.sheets['Expected Goals vs Goals']
# img_expected_goals = Image(plot_path)
# sheet.add_image(img_expected_goals, 'F5')
# wb.save(output_value_path)


# sheet = writer.sheets['Free Kicks Data']
# img_free_kicks = Image(plot_path)
# sheet.add_image(img_free_kicks, 'F5')
# wb.save(output_value_path)


# wb.save(output_value_path)
# print(f"File saved successfully at {output_value_path}")



# team_chosen = 'Bayern München'

# metrics_columns = {
#     "touches_in_opp_box_team": "Touches in Opposition Box",
#     "corner_taken_team": "Corners Taken",
#     "penalty_won_team": "Penalties Won",
#     "expected_goals_team": "Expected Goals",
#     "team_goals_per_match": "Goals per Match",
#     "ontarget_scoring_att_team": "Shots on Target per Match",
#     "possession_won_att_3rd_team": "Possession Won Final 3rd per Match",
#     "big_chance_team": "Big Chances",
#     "won_tackle_team": "Successful Tackles per Match",
#     "clean_sheet_team": "Clean Sheets",
#     "goals_conceded_team_match": "Goals Conceded per Match",
#     "expected_goals_conceded_team": "Expected Goals Conceded",
#     "interception_team": "Interceptions per Match",
#     "penalty_conceded_team": "Penalties Conceded",
#     "saves_team": "Saves per Match",
#     "effective_clearance_team": "Clearances per Match"
# }

# all_teams_data = {}

# for metric_key, column_name in metrics_columns.items():
#     if column_name in df.columns:
#         metric_data = df[['Team', column_name]].drop_duplicates().dropna()

#         ascending = True if "conceded" in column_name.lower() else False
#         sorted_df = metric_data.sort_values(by=column_name, ascending=ascending).reset_index(drop=True)
#         all_teams_data[metric_key] = sorted_df
#     else:
#         print(f"Warning: Metric '{metric_key}' or column '{column_name}' not found. Skipping.")

# fig = make_subplots(
#     rows=4, cols=4, 
#     subplot_titles=list(all_teams_data.keys()), 
#     specs=[[{"type": "bar"}]*4 for _ in range(4)]
# )

# for i, (metric, data) in enumerate(all_teams_data.items()):
#     if data.empty:
#         print(f"Warning: No data available for metric '{metric}'. Skipping.")
#         continue

#     row = i // 4 + 1
#     col = i % 4 + 1

#     colors = ['crimson' if team == team_chosen else 'skyblue' for team in data['Team']]
    
#     fig.add_trace(go.Bar(
#         y=data['Team'], 
#         x=data[data.columns[1]], 
#         marker_color=colors, 
#         orientation='h',
#         text=data[data.columns[1]], 
#         textposition='auto'
#     ), row=row, col=col)

# fig.update_layout(
#     title_text=f"{team_chosen} vs All Teams - Key Metrics",
#     height=2000,
#     width=1500,
#     showlegend=False,
#     title_font=dict(size=20, family='Arial'),
#     margin=dict(l=100, r=50, t=100, b=50),
# )

# try:
#     fig.show()
# except Exception as e:
#     print(f"Error displaying the plot: {e}")
