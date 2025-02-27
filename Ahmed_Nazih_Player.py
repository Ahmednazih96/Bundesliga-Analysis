import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import matplotlib.pyplot as plt 
import seaborn as sns
import plotly.express as px





#from fuzzywuzzy import process 


# print(pd.__version__)
# print(np.__version__)
# import pandas as pd

# df = pd.read_csv("C:\\Users\\user\\Downloads\\bundesliga23_24\\player_expected_goals.csv")


# import pandas as pd

# import pandas as pd

# file_path = r"C:\\Users\\user\\Downloads\\bundesliga23_24\\accurate_cross_team.csv"

# df = pd.read_csv(file_path, encoding='ISO-8859-1')  

# player_name_corrections = {
#     'Niclas FÃ¼llkrug': 'Niclas Füllkrug',
#     'Leroy SanÃ©­': 'Leroy Sané',
#     'Lucas HÃ¶ler': 'Lucas Höler',
#     'Jamal Musiala': 'Jamal Musiala',  
#     'Kevin StÃ¶ger': 'Kevin Stöger',
#     'Mario GÃ¶tze': 'Mario Götze',
#     'Maximilian WÃ¶': 'Maximilian Wö',
#     'Jonas FÃ¶hrenbach': 'Jonas Föhrenbach',
#     'Philipp FÃ¶rster': 'Philipp Förster',
#     'Mert KÃ¶mÃ¼r': 'Mert Kömür',
#     'Grischa PrÃ¶mel': 'Grischa Prömel',
#     'Jannik MÃ¼ller': 'Jannik Müller',
#     'Omar TraorÃ©': 'Omar Traoré',
#     'RogÃ©rio': 'Rogério',
#     'Lukas KÃ¼bler': 'Lukas Kübler',
#     'Niklas SÃ¼le': 'Niklas Süle',
#     'Nicolas HÃ¶fler': 'Nicolas Höfler',
#     'Jan SchÃ¶ppner': 'Jan Schöppner',
#     'Joakim MÃ¦hle': 'Joakim Mæhle',
#     'Marius BÃ¼lter': 'Marius Bülter',
#     'Merlin RÃ¶hl': 'Merlin Röhl',
#     'RaphaÃ«l Guerreiro': 'Raphaël Guerreiro',
#     'Fabian NÃ¼rnberger': 'Fabian Nürnberg',
#     'Christian GÃ¼nter': 'Christian Günter',
#     'JÃ©rome Roussillon': 'Jérôme Roussillon',
#     'Thomas MÃ¼ller': 'Thomas Müller',
#     'Tiago TomÃ¡s': 'Tiago Tomás',
#     'VÃ¡clav Cerny': 'Václav Cerný',
#     'Andreas MÃ¼ller': 'Andreas Müller',
#     'Marcus MÃ¼ller': 'Marcus Müller',
#     'András Schäfer': 'András Schäfer',
#     'Frans KrÃ¤tzig': 'Frans Krätzig',
#     'Chris FÃ¼hrich': 'Chris Führich',
#     'NathanaÃ«l Mbuku': 'Nathanaël Mbuku',
#     'Timo HÃ¼bers': 'Timo Hübers',
#     'Piero HincapiÃ©': 'Piero Hincapié',
#     'AndrÃ¡s SchÃ¤fer': 'András Schäfer',
#     'Maximilian WÃ¶ber': 'Maximilian Wöber',
#     'Kouadio KonÃ©': 'Kouadio Koné',
#     'Kevin MÃ¼ller':'Kevin Muller',
#     'Alexander NÃ¼bel':'Alexander Nobel',
#     'Christian KÃ¼hlwetter':'Christian Kohlwater',
#     'Kjell WÃ¤tjen':'kajil watijin',
#     'PÃ©ter GulÃ¡csi':'Peter Gulacsi',
#     'Marvin SchwÃ¤be':'Marvin Schwabe',
# }

# team_name_corrections = {
#         'Bayern MÃ¼nchen': 'Bayern München',
#     'Borussia MÃ¶nchengladbach': 'Borussia Mönchengladbach',
#     'FC KÃ¶ln': 'FC Köln',
    
# }

# df['Player'] = df['Player'].apply(lambda x: player_name_corrections.get(x, x))
# df['Team'] = df['Team'].apply(lambda x: team_name_corrections.get(x, x))

# try:
#     df.to_csv(file_path,index=False, encoding='ISO-8859-1')  # أو 'utf-16'
#     print("File saved successfully")
#     print(df.head())
# except Exception as e:
#     print(f"Error saving the file: {e}")
# #############################################################################
# import pandas as pd

# file_path = r'C:\Users\user\Downloads\bundesliga23_24\Player\Player_det.xlsx'
# df = pd.read_excel(file_path)


# df.fillna(0, inplace=True)
# print(df.head())
# df.to_excel(r'C:\Users\user\Downloads\bundesliga23_24\Player\modified5_Player_det.xlsx', index=False)

####################################################################################################################

#file_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\New_modified_Player2_det.xlsx"

#df = pd.read_excel(file_path)

#columns_to_replace = ['Expected Assists (xA)', 'Expected Goals (xG)', 'Goals per 90','FotMob Rating','Dribble Success Rate (%)','Successful Dribbles per 90','Fouls Committed per 90','Fouls Won per 90','Clearances per 90','Interceptions per 90','Blocks per 90','Possessions Won Midfield per 90','Possessions Won in Final 3rd per 90','Tackle Success Rate (%)','Tackles per 90','Successful Long Balls (%)','Accurate Long Balls per 90','Pass Success (%)','Accurate Passes per 90','Shot Conversion Rate (%)','Shots per 90','Shot Accuracy (%)','Shot Accuracy (%)','Shots on Target per 90','Big Chances Created','Shot Conversion Rate (%)','Big Chances Missed','Chances Created per 90','Chances Created','Expected Goals on Target (xGOT)',]

#for column in columns_to_replace:
   # mean_value = df[column][df[column] != 0].mean()  # 
   # df[column] = df[column].replace(0, mean_value)  # 

#sorted_columns = df.apply(lambda x: x.sort_values(ascending=False).reset_index(drop=True), axis=0)
#sorted_columns = df.sort_values(by='FotMob Rating', ascending=False)

#rint("\n Sorted all columns done:")
#print(sorted_columns)







# df.insert(0, 'Rank', range(1, len(df) + 1))  



# output_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\Finlay_New_modified_Player2_det.xlsx"
# df.to_excel(output_path, index=False)

# print(f"Save Update: {output_path}")
# print(df.head())
#######################################################################################

#sorted_columns.to_excel(output_path, index=False)


########################################################################################

# df['Rank'] = df['Rank'].sort_values(ascending=True).reset_index(drop=True)

# output_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\Final_modified_Player_det.xlsx"


# sorted_columns = df.drop('Rank', axis=1).apply(lambda x: x.sort_values(ascending=False).reset_index(drop=True), axis=0)

# final_sorted = pd.concat([df[['Rank']], sorted_columns], axis=1)

# final_sorted.to_excel(output_path, index=False)

# print(f"Save Update: {output_path}")

# print(df.head())

#EDA1
##################################################################################
# selected_columns = ['Player', 'Team', 'Actual Goals', 'FotMob Rating', 'Expected Goals (xG)', 'Shot Accuracy (%)', 'Matches']
# df_selected = df[selected_columns]


# missing_values = df.isnull().sum()


# output_value_paith=r"C:\Users\user\OneDrive\Data Curation Progect File\Player\EDA1.xlsx"
# print(missing_values[missing_values > 0]) 
# print(df_selected.head) 
##################################################################################
# sns.histplot(df['Actual Goals'], kde=True, bins=10, color='blue')
# plt.title('Distribution of Goals')
# plt.show()


# sns.kdeplot(df['FotMob Rating'], fill=True, color='green')
# plt.title('FotMob Rating Distribution')
# plt.show()
###################################################################
#print(df.describe())  # avg / mod  
############################################################################
#EDA2



file_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\New_modified_Player2_det.xlsx"

df = pd.read_excel(file_path)

selected_columns = ['Player', 'Team', 'Actual Goals', 'FotMob Rating', 
                    'Expected Goals (xG)', 'Shot Accuracy (%)', 'Matches', 
                    '   ', 'Saves per 90(GK)', 'Total Saves(GK)', 
                    'Clean Sheets(GK)', 'Goals Conceded per 90(GK)', 'Goals Conceded','Chances Created', 'Chances Created per 90', 'Big Chances Created','Big Chances Missed'
                    ,'Blocks per 90','Total Blocks','Interceptions per 90','Total Interceptions','Clearances per 90','Total Clearances','Penalties']
df_selected = df[selected_columns]

missing_values = df.isnull().sum()

output_value_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\Finlay_Player_New_EDA3.xlsx"

df['Performance Category'] = pd.cut(df['FotMob Rating'], bins=[0, 6, 7.5, 10], 
labels=['Low', 'Medium', 'High'])


print(df['Performance Category'].value_counts())


team_goals = df.groupby('Team')['Actual Goals'].sum().sort_values(ascending=False)

print("Total Goals by Team:\n", team_goals)

team_goals.to_excel(output_value_path, sheet_name="Team Goals")

df['Impact Score'] = (
    df['Actual Goals'] * 0.5 +
    df['Actual Assists'] * 0.3 +
    df['FotMob Rating'] * 0.2
)
df['Total_Shots_on_Target']=df['Shots per 90']*df['Matches']
Total_shots_Player=df.groupby('Player')['Total_Shots_on_Target'].sum().sort_values(ascending=False).head(10)

df['Total_Pass_Player']=df['Accurate Passes per 90']*df['Matches']

Total_Passes=df.groupby('Player')['Total_Pass_Player'].sum().sort_values(ascending=False).head(10)


df['Total_long_balls_Player']=df['Accurate Long Balls per 90']*df['Matches']
Total_long_ball=df.groupby('Player')['Total_long_balls_Player'].sum().sort_values(ascending=False).head(10)


df['Total_Tackles_Player']=df['Tackles per 90']*df['Matches']
Total_Tackles=df.groupby('Player')['Total_Tackles_Player'].sum().sort_values(ascending=False).head(10)


df['Total_Possessions_Won_Midfield_Player']=df['Possessions Won Midfield per 90']*df['Matches']
Total_Possessions_Won_Midfield=df.groupby('Player')['Total_Possessions_Won_Midfield_Player'].sum().sort_values(ascending=False).head(10)


df['Total_Player_Possessions_Won_in_Final_3rd']=df['Possessions Won in Final 3rd per 90']*df['Matches']
Total_Possessions_Won_in_Final_3rd=df.groupby('Player')['Total_Player_Possessions_Won_in_Final_3rd'].sum().sort_values(ascending=False).head(10)


df['Total_Player_Successful_Dribbles']=df['Successful Dribbles per 90']*df['Matches']
Total_Successful_Dribbles=df.groupby('Player')['Total_Player_Successful_Dribbles'].sum().sort_values(ascending=False).head(10)


#Successful Dribbles per 90



#print(Total_shots_Player)
top_scored_plantes=df.sort_values(by='Penalties', ascending=False).head(10)
Top_Player_goals = df.groupby('Player')['Actual Goals'].sum().sort_values(ascending=False).head(15)
#print("Top player Goals",Top_Player_goals)
top_assists = df.sort_values(by='Actual Assists', ascending=False).head(15)

top_players = df.sort_values(by='Impact Score', ascending=False).head(10)

assist_data = df[['Player', 'Team', 'Matches', 'Expected Assists (xA)', 'Actual Assists']]
chances_created_columns = ['Player', 'Team', 'Chances Created', 'Chances Created per 90', 'Big Chances Created']
chances_data = df[chances_created_columns]

fouls_cards_columns = ['Player', 'Team', 'Fouls Won per 90', 'Fouls Committed per 90', 'Red Cards', 'Yellow Cards','Matches']
fouls_cards_data = df[fouls_cards_columns]

Total_Top_Yellow_Card=df.groupby('Player')['Yellow Cards'].sum().sort_values(ascending=False).head(10)


Total_Top_Red_Card=df.groupby('Player')['Red Cards'].sum().sort_values(ascending=False).head(10)


df['top_fouls_won']=df['Fouls Won per 90']*df['Matches']
Total_top_fouls_won=df.groupby('Player')['top_fouls_won'].sum().sort_values(ascending=False).head(10)

#top_fouls_won = fouls_cards_data.sort_values(by='Fouls Won per 90', ascending=False).head(10)
print("Top 5 Players by Fouls Won per 90:")
#print(top_fouls_won)

df['top_Committed_won']=df['Fouls Committed per 90']*df['Matches']
Total_top_Committed_won=df.groupby('Player')['top_Committed_won'].sum().sort_values(ascending=False).head(10)

#top_fouls_committed = fouls_cards_data.sort_values(by='Fouls Committed per 90', ascending=False).head(10)
print("Top 5 Players by Fouls Committed per 90:")
#print(top_fouls_committed)

# most_rough_players = pd.concat([
#     Total_top_fouls_won, 
#     Total_top_Committed_won, 
#     Total_Top_Red_Cards, 
#     Total_Top_Yellow_Cards
# ], ignore_index=True).drop_duplicates()

#print("Most Rough Players (Top Fouls Committed and Fouls Won):")
#print(most_rough_players)
top_all_Clearances=df_selected
top_blocks = top_all_Clearances.sort_values(by='Blocks per 90', ascending=False).head(10)
top_interceptions = top_all_Clearances.sort_values(by='Interceptions per 90', ascending=False).head(10)
top_clearances = top_all_Clearances.sort_values(by='Clearances per 90', ascending=False).head(10)
top_total_blocks = top_all_Clearances.sort_values(by='Total Blocks', ascending=False).head(10)
top_total_interceptions = top_all_Clearances.sort_values(by='Total Interceptions', ascending=False).head(10)
top_total_clearances = top_all_Clearances.sort_values(by='Total Clearances', ascending=False).head(10)

defensive_stats = pd.concat([top_blocks, top_total_blocks, top_interceptions, 
                             top_total_interceptions, top_clearances, top_total_clearances]).head(10)


#print(chances_data.head())
top_chances_created = chances_data.sort_values(by='Chances Created', ascending=False).head(10)
print("Top 10 Players by Chances Created:")
#print(top_chances_created)

top_big_chances_created = chances_data.sort_values(by='Big Chances Created', ascending=False).head(10)
print("Top 10 Players by Big Chances Created:")
print(top_big_chances_created)




assist_data_sorted = assist_data.sort_values(by='Actual Assists', ascending=False)



df['Goal Improvement'] = df['Expected Goals (xG)'] - df['Actual Goals']
underperforming_players = df[df['Goal Improvement'] > 0]

goalkeeper_columns = ['Player', 'Team', 'Saves per 90(GK)', 'Total Saves(GK)', 
                      'Clean Sheets(GK)', 'Goals Conceded per 90(GK)', 'Goals Conceded']
goalkeepers_data = df[goalkeeper_columns]

top_clean_sheets = goalkeepers_data.sort_values(by='Clean Sheets(GK)', ascending=False).head(18)
top_saves_per_90 = goalkeepers_data.sort_values(by='Saves per 90(GK)', ascending=False).head(18)


if os.path.exists(output_value_path):
    os.remove(output_value_path)

with pd.ExcelWriter(output_value_path, engine='openpyxl') as writer:
    df_selected.to_excel(writer, sheet_name='Selected Data', index=False)
    missing_values[missing_values > 0].to_frame(name='Missing Values').to_excel(writer, sheet_name='Missing Values')
    df.describe().to_excel(writer, sheet_name='Descriptive Stats')
    df['Performance Category'].value_counts().to_frame(name='Performance Counts').to_excel(writer, sheet_name='Performance Counts')
    team_goals.to_excel(writer, sheet_name='Team Goals')
    top_players.to_excel(writer, sheet_name="Top Players", index=False)
    underperforming_players.to_excel(writer, sheet_name="Underperforming Players", index=False)
    goalkeepers_data.to_excel(writer, sheet_name='Goalkeepers Data', index=False)
    top_clean_sheets.to_excel(writer, sheet_name='Top Clean Sheets', index=False)
    top_saves_per_90.to_excel(writer, sheet_name='Top Saves per 90', index=False)
    top_assists.to_excel(writer, sheet_name='Top Assists', index=False)
    assist_data_sorted.to_excel(writer, sheet_name="Assist Data", index=False)
    chances_data.to_excel(writer, sheet_name="Chances Created Data", index=False)
    top_chances_created.to_excel(writer, sheet_name="Top Players by Chances Created", index=False)
    top_big_chances_created.to_excel(writer, sheet_name="Top Players by Big Chances Created", index=False)
    fouls_cards_data.to_excel(writer, sheet_name="Fouls and Cards Data", index=False)
    Total_top_fouls_won.to_excel(writer, sheet_name="Top Players by Fouls Won per 90")
    Total_top_Committed_won.to_excel(writer, sheet_name="Top Players by Fouls Committed per 90")
    #most_rough_players.to_excel(writer, sheet_name="Most Rough Players",index=False)
    defensive_stats.to_excel(writer, sheet_name="Defensive Stats", index=False)
    Top_Player_goals.to_excel(writer,sheet_name="Top Player goals" )
    top_scored_plantes.to_excel(writer,sheet_name="top scored plantes" )
    Total_shots_Player.to_excel(writer,sheet_name="Total_shots_Player_on_Target")
    Total_Passes.to_excel(writer,sheet_name="Total_Pass_Player")
    Total_long_ball.to_excel(writer,sheet_name="Top_Crosses Ball")
    Total_Tackles.to_excel(writer,sheet_name="Top_Total_Tackles_Player")
    Total_Possessions_Won_Midfield.to_excel(writer,sheet_name="Total_Possessions_Won_Midfield_Player")
    Total_Possessions_Won_in_Final_3rd.to_excel(writer,sheet_name="Total_Player_Possessions_Won_in_Final_3rd")
    Total_Successful_Dribbles.to_excel(writer,sheet_name="Most_Player_Successful_Dribbles")
    Total_Top_Yellow_Card.to_excel(writer,sheet_name="Total_Top_Yellow_Card_player")
    Total_Top_Red_Card.to_excel(writer,sheet_name="Total_Top_Red_Card")
    top_total_interceptions.to_excel(writer,sheet_name="Top Total interceptions")
    top_total_blocks.to_excel(writer,sheet_name="Top Total Blocks")
    







plt.figure(figsize=(10, 6))
sns.histplot(df['Actual Goals'], kde=True, bins=10, color='blue')
plt.title('Distribution of Goals')
plt.xlabel('Actual Goals')
plt.ylabel('Frequency')
plt.tight_layout()
plot_path_goals = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\histogram_goals.png"
plt.savefig(plot_path_goals)
plt.close()

plt.figure(figsize=(10, 6))
sns.kdeplot(df['FotMob Rating'], fill=True, color='green')
plt.title('FotMob Rating Distribution')
plt.xlabel('FotMob Rating')
plt.ylabel('Density')
plt.tight_layout()
plot_path_rating = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\kdeplot_rating.png"
plt.savefig(plot_path_rating)
plt.close()

plt.figure(figsize=(10, 6))
sns.scatterplot(data=df, x='Actual Goals', y='FotMob Rating', hue='Team', size='Matches')
plt.title('Goals vs FotMob Rating')
plt.tight_layout()
scatterplot_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\scatterplot_goals_rating.png"
plt.savefig(scatterplot_path)
plt.close()

numeric_df = df.select_dtypes(include=['float64', 'int64'])
correlation_matrix = numeric_df.corr()


plt.figure(figsize=(12, 10))
sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm')
plt.title('Correlation Heatmap')
plt.tight_layout()
heatmap_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\correlation_heatmap.png"
plt.savefig(heatmap_path)
plt.close()

plt.figure(figsize=(10, 6))
sns.regplot(data=df, x='Shot Accuracy (%)', y='Actual Goals', line_kws={'color': 'red'})
plt.title('Shot Accuracy vs Goals')
plt.tight_layout()
regplot_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\regplot_accuracy_goals.png"
plt.savefig(regplot_path)
plt.close()

df['Performance Category'].value_counts().plot(kind='bar', color=['red', 'orange', 'green'])
plt.title('Player Performance Categories')
plt.xlabel('Performance Category')
plt.ylabel('Number of Players')
plt.tight_layout()
performance_chart_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\Performance_Categories.png"
plt.savefig(performance_chart_path)
plt.close()

team_goals.plot(kind='bar', figsize=(10, 6), color='blue')
plt.title('Total Goals by Team')
plt.xlabel('Team')
plt.ylabel('Goals')
plt.tight_layout()
team_goals_chart_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\Team_Goals.png"
plt.savefig(team_goals_chart_path)
plt.close()

plt.figure(figsize=(10, 6))
sns.barplot(data=top_players, x='Impact Score', y='Player', palette='viridis')
plt.title('Top 10 Most Influential Players')
plt.xlabel('Impact Score')
plt.ylabel('Player')
plt.tight_layout()
top_players_chart_path = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\top_players.png"
plt.savefig(top_players_chart_path)
plt.close()

plt.figure(figsize=(10, 6))
sns.scatterplot(data=goalkeepers_data, x='Total Saves(GK)', y='Goals Conceded', color='red')
plt.title('Total Saves vs Goals Conceded')
plt.xlabel('Total Saves (GK)')
plt.ylabel('Goals Conceded')
plt.tight_layout()
saves_vs_goals_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\saves_vs_goals.png"
plt.savefig(saves_vs_goals_plot)
plt.close()

sns.barplot(data=top_players, x='Impact Score', y='Player', palette='viridis')
plt.title('Top 10 Most Influential Players')
plt.xlabel('Impact Score')
plt.ylabel('Player')
top_players_chart = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\top_players.png"
plt.savefig(top_players_chart)
plt.close()

plt.figure(figsize=(10, 6))
sns.histplot(goalkeepers_data['Saves per 90(GK)'], kde=True, color='blue')
plt.title('Distribution of Saves per 90 (GK)')
plt.xlabel('Saves per 90')
plt.ylabel('Frequency')
plt.tight_layout()
saves_per_90_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\saves_per_90.png"
plt.savefig(saves_per_90_plot)
plt.close()

plt.figure(figsize=(10, 6))
sns.histplot(goalkeepers_data['Clean Sheets(GK)'], kde=True, color='green')
plt.title('Distribution of Clean Sheets (GK)')
plt.xlabel('Clean Sheets')
plt.ylabel('Frequency')
plt.tight_layout()
clean_sheets_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\clean_sheets.png"
plt.savefig(clean_sheets_plot)
plt.close()

plt.figure(figsize=(10, 6))
sns.scatterplot(data=goalkeepers_data, x='Total Saves(GK)', y='Goals Conceded', color='red')
plt.title('Total Saves vs Goals Conceded')
plt.xlabel('Total Saves (GK)')
plt.ylabel('Goals Conceded')
plt.tight_layout()
saves_vs_goals_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\saves_vs_goals.png"
plt.savefig(saves_vs_goals_plot)
plt.close()

plt.figure(figsize=(12, 8))
sns.barplot(data=top_assists, x='Actual Assists', y='Player', palette='coolwarm')
plt.title('Top 15 Players by Assists')
plt.xlabel('Actual Assists')
plt.ylabel('Player')
plt.tight_layout()
top_assists_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\top_assists.png"
plt.savefig(top_assists_plot)
plt.close()

plt.figure(figsize=(12, 8))
sns.scatterplot(data=df, x='Expected Assists (xA)', y='Actual Assists', hue='Team', palette='tab10', size='Matches', sizes=(50, 200))
plt.title('Expected vs Actual Assists')
plt.xlabel('Expected Assists (xA)')
plt.ylabel('Actual Assists')
plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
plt.tight_layout()
expected_vs_actual_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\expected_vs_actual_assists.png"
plt.savefig(expected_vs_actual_plot)
plt.close()

plt.figure(figsize=(10, 6))
sns.histplot(chances_data['Chances Created'], kde=True, color='purple')
plt.title('Distribution of Chances Created')
plt.xlabel('Chances Created')
plt.ylabel('Frequency')
plt.tight_layout()
chances_created_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\chances_created.png"
plt.savefig(chances_created_plot)
plt.close()

plt.figure(figsize=(10, 6))
sns.histplot(chances_data['Big Chances Created'], kde=True, color='orange')
plt.title('Distribution of Big Chances Created')
plt.xlabel('Big Chances Created')
plt.ylabel('Frequency')
plt.tight_layout()
big_chances_created_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\big_chances_created.png"
plt.savefig(big_chances_created_plot)
plt.close()

plt.figure(figsize=(12, 6))
top_players = pd.concat([top_blocks[['Player', 'Blocks per 90']], 
                         top_interceptions[['Player', 'Interceptions per 90']], 
                         top_clearances[['Player', 'Clearances per 90']]])

top_players_melted = top_players.melt(id_vars="Player", var_name="Metric", value_name="Value")

top_players_melted['Value'] = pd.to_numeric(top_players_melted['Value'], errors='coerce')

sns.barplot(data=top_players_melted, x='Player', y='Value', hue='Metric')
plt.title('Top 10 Players Comparison - Blocks, Interceptions, and Clearances')
plt.xlabel('Player')
plt.ylabel('Value per 90')
plt.xticks(rotation=90)
plt.tight_layout()
top_players_melted_plot = r"C:\Users\user\OneDrive\Data Curation Progect File\Player\top_players_melted_plot.png"
plt.savefig(top_players_melted_plot)
plt.close()


workbook = load_workbook(output_value_path)
sheet = workbook.create_sheet('Visualizations')

sheet2=workbook.create_sheet('chartsheet')

sheet3=workbook.create_sheet('GKSheet')

sheet4=workbook.create_sheet('ChancesChart')



img_goals = Image(plot_path_goals)
sheet.add_image(img_goals, 'A1')

img_rating = Image(plot_path_rating)
sheet.add_image(img_rating, 'A20')
img_scatterplot = Image(scatterplot_path)
sheet.add_image(img_scatterplot, 'A40')

img_heatmap = Image(heatmap_path)
sheet.add_image(img_heatmap, 'A90')

img_regplot = Image(regplot_path)
sheet.add_image(img_regplot, 'A120')
workbook.save(output_value_path)

img_team_goals = Image(performance_chart_path)
sheet2.add_image(img_team_goals, 'A2')
workbook.save(output_value_path)

img_performance_chart = Image(team_goals_chart_path)
sheet2.add_image(img_performance_chart, 'A70')
workbook.save(output_value_path)

img_top_players = Image(top_players_chart_path)
sheet2.add_image(img_top_players, 'A110')
workbook.save(output_value_path)

img_saves_per_90 = Image(saves_per_90_plot)
sheet3.add_image(img_saves_per_90, 'A2')
workbook.save(output_value_path)


img_clean_sheets = Image(clean_sheets_plot)
sheet3.add_image(img_clean_sheets, 'A50')
workbook.save(output_value_path)


img_saves_vs_goals_plot = Image(saves_vs_goals_plot)
sheet3.add_image(img_saves_vs_goals_plot, 'A100')
workbook.save(output_value_path)

img_top_assists = Image(top_assists_plot)
sheet2.add_image(img_top_assists, 'A180')   
workbook.save(output_value_path)

img_expected_vs_actual = Image(expected_vs_actual_plot)
sheet2.add_image(img_expected_vs_actual, 'A150') 
workbook.save(output_value_path)

img_chances_created_plot  = Image(chances_created_plot )
sheet4.add_image(img_chances_created_plot, 'A2') 
workbook.save(output_value_path)

img_big_chances_created_plo  = Image(big_chances_created_plot )
sheet4.add_image(img_big_chances_created_plo, 'A40') 
workbook.save(output_value_path)

img_top_players_melted_plot  = Image(top_players_melted_plot )
sheet4.add_image(img_top_players_melted_plot, 'A80') 
workbook.save(output_value_path)

print(f"Data and visualizations successfully saved to {output_value_path}")
###################################################################################



















